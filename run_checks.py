#!/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import requests
import yaml
from requests.auth import HTTPBasicAuth
import pdb
import logging
import sys
import os
import openpyxl
from datetime import datetime
from datetime import timedelta

# employment grace period in days
EMPLOYMENT_GRACE_PERIOD_DAYS = 30


def download_spreadsheet(config, args):
    """
    Download a spreadsheet from a given URL using basic authentication.
    """

    logging.debug("Starting download_spreadsheet function...")

    # Download the spreadsheet
    logging.info("Downloading spreadsheet...")
    logging.debug(f"Using URL: {config['base_url']}/remote.php/dav/files/{config['username']}/{config['remote_file_path']}")
    response = requests.get(f"{config['base_url']}/remote.php/dav/files/{config['username']}/{config['remote_file_path']}", auth=HTTPBasicAuth(config['username'], config['password']))

    # Check if the request was successful
    if response.status_code == 200:
        with open(args.file, 'wb') as f:
            f.write(response.content)
        logging.info("Spreadsheet downloaded successfully.")
    else:
        logging.error("Failed to download the spreadsheet.")




def run_checks(file_path):
    """
    Run checks on the downloaded spreadsheet.
    """

    logging.debug("Starting run_checks function...")

    # Load the staff spreadsheet
    logging.info("Loading spreadsheet...")
    try:
        workbook = openpyxl.load_workbook(file_path)
        staff_sheet = workbook['Staff']
    except Exception as e:
        logging.error(f"Failed to load spreadsheet: {e}")
        sys.exit(1)

    # load the exception spreadsheet
    logging.info("Loading exception spreadsheet...")
    try:
        exception_sheet = workbook['Exceptions']
    except Exception as e:
        logging.error(f"Failed to load exception spreadsheet: {e}")
        sys.exit(1)
    exceptions = parse_exceptions(exception_sheet)

    # create header to column mapping
    logging.debug("Creating header to column mapping...")
    headers = staff_sheet[1]  # Assuming the first row contains headers
    header_mapping = {}
    for col in range(0, len(headers)):
        header_mapping[headers[col].value] = col
    logging.debug(f"Header mapping: {header_mapping}")

    # initialize a dictionary to store warnings
    warnings = {}

    # for each row in the spreadsheet
    logging.info("Running checks on the spreadsheet...")
    for row in staff_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        logging.debug(f"Processing row: {row}")

        # check if first column is empty, i.e. end of the list
        if row[0] is None:
            logging.debug("First column is empty, ending processing.")
            break

        # readability
        user_mail = row[header_mapping['nbis mail']]
        user_warnings = set()

        # get execeptions if any
        if user_mail in exceptions:
            user_exceptions = exceptions[user_mail]
        else:
            user_exceptions = set()
        
        # check if the employee has been terminated more than n days ago
        if not check_employee_active(row, header_mapping, EMPLOYMENT_GRACE_PERIOD_DAYS=EMPLOYMENT_GRACE_PERIOD_DAYS):

            # check if the email is still active
            if row[header_mapping['mail active']] == "yes":
                logging.debug(f"Email is still active: {user_mail}")
                # check if the user has exceptions
                if "mail active" in user_exceptions:
                    logging.debug(f"User {user_mail} has exception for mail active.")
                else:
                    user_warnings.add("mail active")

            # check if github is still active
            if row[header_mapping['github active']] == "yes":
                logging.debug(f"Github is still active: {user_mail}")
                # check if the user has exceptions
                if "github active" in user_exceptions:
                    logging.debug(f"User {user_mail} has exception for github active.")
                else:
                    user_warnings.add("github active")

            # check if confluence is still active
            if row[header_mapping['confluence active']] == "yes":
                logging.debug(f"Confluence is still active: {user_mail}")
                # check if the user has exceptions
                if "confluence active" in user_exceptions:
                    logging.debug(f"User {user_mail} has exception for confluence active.")
                else:
                    user_warnings.add("confluence active")


            # check if there were any warnings
            if len(user_warnings) > 0:
                logging.debug(f"User {user_mail} has warnings: {user_warnings}")
                warnings[user_mail] = user_warnings


    # check if there were any warnings
    if len(warnings) > 0:
        logging.warning("Warnings found:")
        for user_mail, user_warnings in warnings.items():
            logging.warning(f"User {user_mail} has warnings: {user_warnings}")

# function to check if a employee is still active
def check_employee_active(row, header_mapping, EMPLOYMENT_GRACE_PERIOD_DAYS=EMPLOYMENT_GRACE_PERIOD_DAYS):
    """
    Check if an employee is still active.
    """

    logging.debug("Starting check_employee_active function...")

    # check if the employee has been terminated
    if row[header_mapping['employment end']] is not None:
        employment_end_date = row[header_mapping['employment end']]

        # check if the date is more than n days ago
        date_grace_period = datetime.now() - timedelta(days=(EMPLOYMENT_GRACE_PERIOD_DAYS))
        if employment_end_date < date_grace_period:
            logging.debug(f"Employee terminated more than {EMPLOYMENT_GRACE_PERIOD_DAYS} days ago: {row[header_mapping['nbis mail']]}")
            return False

    return True


# function to parse the exceptions spreadsheet
def parse_exceptions(exception_sheet):
    """
    Parse the exceptions spreadsheet.
    """

    logging.debug("Starting parse_exceptions function...")

    # create a dictionary to store exceptions
    exceptions = {}
    #pdb.set_trace()

    # for each row in the spreadsheet
    for row in exception_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        logging.debug(f"Processing exception row: {row}")

        # check if first column is empty, i.e. end of the list
        if row[0] is None:
            logging.debug("First column is empty, ending processing.")
            break

        # readabilty
        user_mail = row[0]
        try:
            user_exceptions = row[1].split(",")
        except Exception as e:
            user_exceptions = []

        # add the exception to the dictionary
        logging.debug(f"Adding exception for user: {user_mail} with exception(s): {user_exceptions}")
        for user_exception in user_exceptions:
            logging.debug(f"Processing user exception: {user_exception.strip()}")

            # check if the user already has exceptions
            if user_mail not in exceptions:
                exceptions[user_mail] = set()

            # add the exception to the set
            logging.debug(f"Adding exception: {user_exception.strip()} to user: {user_mail}")
            exceptions[user_mail].add(user_exception.strip())

    return exceptions

if __name__ == "__main__":

    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Process some integers.')
    parser.add_argument('-c', '--config', type=str, help='Path to YAML configuration file', required=True)
    parser.add_argument('-d', '--debug', action='store_true', help='Enable debug mode')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose mode')
    parser.add_argument('-f', '--file', type=str, help='Path to where the stafflist will be downloaded (default: /tmp/)', default='/tmp/nbis_staff.xlsx')
    parser.add_argument('-k', '--keep', action='store_true', help='Keep the downloaded file after processing')
    args = parser.parse_args()

    # Set up logging
    logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

    # Enable verbose mode if specified
    if args.verbose:
        logging.getLogger().setLevel(logging.INFO)
        logging.info("Verbose mode enabled.")

    # Enable debug mode if specified
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.debug("Debug mode enabled.")

    # Read in user credentials and other configs from the YAML file
    logging.debug("Reading configuration file...")
    with open(args.config, 'r') as file:
        config = yaml.safe_load(file)

    # download the spreadsheet
    download_spreadsheet(config, args)

    # run checks on the downloaded file
    run_checks(args.file)

    # remove the downloaded file if not keeping it
    logging.debug("Checking if the file should be removed...")
    if not args.keep:
        logging.debug("Removing the downloaded file...")
        try:
            os.remove(args.file)
            logging.info(f"Removed downloaded file: {args.file}")
        except Exception as e:
            logging.error(f"Failed to remove file: {e}")
            sys.exit(1)

