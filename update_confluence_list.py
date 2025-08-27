#!/bin/env python3
# -*- coding: utf-8 -*-

import argparse
#import requests
import yaml
from requests.auth import HTTPBasicAuth
import pdb
import logging
import sys
import os
import openpyxl
from datetime import datetime
from datetime import timedelta
from Nextcloud import Nextcloud
from Confluence import Confluence
from pprint import pprint

# update the confluence staff list


if __name__ == "__main__":

    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Process some integers.')
    parser.add_argument('-c', '--config', type=str, help='Path to YAML configuration file', required=True)
    parser.add_argument('-d', '--debug', action='store_true', help='Enable debug mode')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose mode')
    parser.add_argument('-f', '--file', type=str, help='Path to where the stafflist will be downloaded (default: /tmp/)', default='/tmp/nbis_staff.xlsx')
    parser.add_argument('-k', '--keep', action='store_true', help='Keep the downloaded file after processing')
    args = parser.parse_args()

    # Set up logging
    logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

    # Enable verbose mode if specified
    if args.verbose:
        logging.getLogger().setLevel(logging.INFO)
        logging.info("Verbose mode enabled.")

    # Enable debug mode if specified
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.debug("Debug mode enabled.")

    # Read in user credentials and other configs from the YAML file
    logging.debug("Reading configuration file...")
    with open(args.config, 'r') as file:
        config = yaml.safe_load(file)

    # download the spreadsheet
    nextcloud = Nextcloud(config, args)
    nextcloud.download_spreadsheet()

    # Check if the file exists
    logging.debug(f"Checking if the file exists at {args.file}...")
    if not os.path.exists(args.file):
        logging.error(f"File does not exist: {args.file}")
        sys.exit(1)
    logging.info(f"File exists: {args.file}")

    # Load the spreadsheet
    logging.debug("Loading the spreadsheet...")
    try:
        wb = openpyxl.load_workbook(args.file)
        ws = wb.active
    except Exception as e:
        logging.error(f"Failed to load the spreadsheet: {e}")
        sys.exit(1)

    logging.info("Spreadsheet loaded successfully.")

    # get spreadsheet modification time
    mod_time = datetime.fromtimestamp(os.path.getmtime(args.file))
    logging.info(f"Spreadsheet last modified time: {mod_time.isoformat()}")


    # initialize the html table
    logging.debug("Initializing the HTML table for staff list...")
    html_table = f"""
    <table>
        <tr>
            <th><strong>Name</strong></th>
            <th><strong>Unit</strong></th>
            <th><strong>Organization</strong></th>
            <th><strong>Role</strong></th>
        </tr>
    """

    # get the column headers and make a dictionary of the column names
    logging.debug("Getting column headers...")
    headers = [cell.value.lower() if cell.value is not None else cell.value for cell in ws[1]]
    header_dict = {header: index for index, header in enumerate(headers)}
    logging.debug(f"Column headers: {header_dict}")

    # Loop through the rows in the spreadsheet and create the HTML table
    logging.debug("Processing the spreadsheet...")
    staff_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
    
        html_table += f"""
        <tr>
            <td>{row[header_dict['name']]}</td>
            <td>{row[header_dict['unit']]}</td>
            <td>{row[header_dict['organization']]}</td>
            <td>{row[header_dict['role']]}</td>
        </tr>"""

    html_table += f"""\n    </table>
This list is generated daily from the central master staff list in Data Center's NextCloud instance.
Staff list document updated:&nbsp;&nbsp;&nbsp;{mod_time.strftime("%Y-%m-%d %H:%M")}
This page rendered on:&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{datetime.now().strftime("%Y-%m-%d %H:%M")}
"""


    confluence = Confluence(config, args)
    logging.debug("Updating the Confluence staff list page...")
    updated = confluence.update_staff_list_page(config['confluence']['space_key'], config['confluence']['page_id'], html_content=html_table)

    # remove the downloaded file if not keeping it
    logging.debug("Checking if the file should be removed...")
    if not args.keep:
        logging.debug("Removing the downloaded file...")
        try:
            os.remove(args.file)
            logging.info(f"Removed downloaded file: {args.file}")
        except Exception as e:
            logging.error(f"Failed to remove file: {e}")
            sys.exit(2)
    else:
        logging.info("Keeping the downloaded file as per user request.")

